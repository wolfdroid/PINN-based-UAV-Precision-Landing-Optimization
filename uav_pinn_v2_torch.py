"""
=============================================================================
  UAV Drone Precision Landing — PINN Control Strategy v2.0 (PyTorch + CUDA)
  Full-featured implementation with torch.autograd and GPU acceleration.
=============================================================================

Requirements:
    pip install torch numpy matplotlib

Usage:
    python uav_pinn_v2_torch.py                  # Auto-detect GPU
    python uav_pinn_v2_torch.py --device cpu      # Force CPU
    python uav_pinn_v2_torch.py --device cuda      # Force GPU
    python uav_pinn_v2_torch.py --epochs 8000      # Custom epochs

Changes implemented (identical to NumPy v2, now with proper autograd):
-----------------------------------------------------------
[1]  TIME-VARYING CONTROL PROFILE via τ-conditioned network
[2]  YAW ANGLE EXPLICITLY FIXED AT ZERO (ψ = 0)
[3]  HORIZONTAL VELOCITY PENALTY AT LANDING (ẋ² + ẏ²)
[4]  GROUND COLLISION PENALTY: max(0, −z)² at collocation points
[5]  CONTROL FEASIBILITY BOUNDS via tanh/sigmoid
[6]  GROUND EFFECT OFF by default (toggleable flag)
[7]  ADAPTIVE OUTPUT SCALING from config ranges
[8]  COSINE ANNEALING LR (Adam) + L-BFGS precision phase
[9]  EXPLICIT INITIAL CONDITIONS: (0, 0, z₀) with zero velocity
[10] COMPREHENSIVE LOSS LOGGING (per-component)
[11] MULTI-SCENARIO EVALUATION (4 test cases)
[12] PRODUCTION VISUALIZATION (6-panel dark dashboard)
[13] EXACT AUTOGRAD PDE RESIDUAL: dS/dτ = T · f(S, u)
[14] RK4 NUMERICAL VERIFICATION (independent forward integration)
[15] ASYMMETRIC DRAG (k_dx, k_dy, k_dz)

Key improvements over NumPy version:
  - torch.autograd gives EXACT gradients (no finite-difference noise)
  - Full gradient flow through physics equations AND control constraints
  - L-BFGS second phase with strong Wolfe line search
  - CUDA acceleration for all tensor operations
  - Mixed-precision ready (can be enabled for A100/H100)
"""

import argparse
import math
import os
import platform
import random
import time
from dataclasses import dataclass
from typing import Dict, List, Tuple

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import numpy as np
import torch
import torch.nn as nn

try:
    import psutil
    HAS_PSUTIL = True
except ImportError:
    HAS_PSUTIL = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ═══════════════════════════════════════════════════════════════════════════
#  1. CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════

@dataclass
class UAVParams:
    """Physical parameters of the quadrotor UAV."""
    m: float = 1.5              # Mass (kg)
    g: float = 9.81             # Gravity (m/s²)
    k_dx: float = 0.25          # [15] Per-axis drag coefficients
    k_dy: float = 0.25
    k_dz: float = 0.30          # Vertical drag typically higher
    z0: float = 10.0            # Initial altitude (m)
    yaw: float = 0.0            # [2] Yaw fixed at 0°
    theta_max_deg: float = 30.0 # Max pitch angle (degrees)
    phi_max_deg: float = 30.0   # Max roll angle (degrees)
    thrust_min_ratio: float = 0.5
    thrust_max_ratio: float = 2.0
    T_min: float = 2.0          # Min flight time (s)
    T_max: float = 8.0          # Max flight time (s)
    enable_ground_effect: bool = False  # [6] Off by default
    R_rotor: float = 0.25       # Rotor radius (used only if ground effect ON)


@dataclass
class TrainConfig:
    """Training hyperparameters."""
    adam_epochs: int = 4000      # Phase 1: Adam global search
    lbfgs_epochs: int = 500      # Phase 2: L-BFGS precision tuning
    batch_size: int = 512        # Larger batch → better gradient estimates
    n_colloc: int = 80           # Collocation points for PDE residual
    lr: float = 1e-3             # Initial learning rate (Adam)
    print_every: int = 500       # Log interval
    target_xy_range: float = 8.0 # Task sampling range for (xf, yf)
    wind_range: float = 3.0      # Task sampling range for (vwx, vwy)

    # Loss weights
    w_pde: float = 2.0           # PDE physics residual
    w_ic: float = 40.0           # Initial condition
    w_bc: float = 40.0           # Boundary condition (landing)
    w_ground: float = 10.0       # [4] Ground collision penalty
    w_vel_hz: float = 15.0       # [3] Horizontal velocity at landing


# ═══════════════════════════════════════════════════════════════════════════
#  2. DEVICE SELECTION (CUDA / MPS / CPU)
# ═══════════════════════════════════════════════════════════════════════════

def select_device(requested: str = "auto") -> torch.device:
    """Select the best available device."""
    if requested == "auto":
        if torch.cuda.is_available():
            dev = torch.device("cuda")
            name = torch.cuda.get_device_name(0)
            mem = torch.cuda.get_device_properties(0).total_memory / 1e9
            print(f"  Device: CUDA — {name} ({mem:.1f} GB)")
        elif hasattr(torch.backends, "mps") and torch.backends.mps.is_available():
            dev = torch.device("mps")
            print("  Device: MPS (Apple Silicon)")
        else:
            dev = torch.device("cpu")
            print("  Device: CPU")
    else:
        dev = torch.device(requested)
        print(f"  Device: {requested}")
    return dev


# ═══════════════════════════════════════════════════════════════════════════
#  3. PINN NETWORK (PyTorch)
# ═══════════════════════════════════════════════════════════════════════════

class PINNNetwork(nn.Module):
    """
    Physics-Informed Neural Network for UAV landing control.

    Input:  τ ∈ [0,1] (normalized time) + task [xf, yf, vwx, vwy]  → (B, N, 5)
    Output: states [x, y, z, vx, vy, vz], controls [θ, φ, F], flight time T

    Architecture: 5 → 256 → 256 → 256 → 256 → 10  (4 hidden layers, tanh)
    Skip connections on layers 2→3 and 3→4 for better gradient flow.
    """

    def __init__(self, params: UAVParams, cfg: TrainConfig, hidden_dim: int = 256):
        super().__init__()
        self.params = params

        # 4-layer MLP with skip connections
        self.fc0 = nn.Linear(5, hidden_dim)
        self.fc1 = nn.Linear(hidden_dim, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.fc3 = nn.Linear(hidden_dim, hidden_dim)
        self.fc_out = nn.Linear(hidden_dim, 10)

        # Xavier initialization
        for m in self.modules():
            if isinstance(m, nn.Linear):
                nn.init.xavier_normal_(m.weight)
                nn.init.zeros_(m.bias)

        # Physical bounds
        self.theta_max = math.radians(params.theta_max_deg)
        self.phi_max = math.radians(params.phi_max_deg)
        self.F_min = params.thrust_min_ratio * params.m * params.g
        self.F_max = params.thrust_max_ratio * params.m * params.g

        # [7] Adaptive output scaling
        self.xy_scale = cfg.target_xy_range * 1.5
        self.z_scale = params.z0
        self.v_scale = cfg.target_xy_range

    def forward(self, tau: torch.Tensor, task: torch.Tensor
                ) -> Tuple[torch.Tensor, torch.Tensor, torch.Tensor]:
        """
        Args:
            tau:  (B, N, 1) normalized time
            task: (B, N, 4) task parameters [xf, yf, vwx, vwy]
        Returns:
            states: (B, N, 6)  [x, y, z, vx, vy, vz]
            ctrls:  (B, N, 3)  [theta, phi, F]
            T:      (B, N, 1)  flight time
        """
        inp = torch.cat([tau, task], dim=-1)  # (B, N, 5)

        h = torch.tanh(self.fc0(inp))
        h1 = torch.tanh(self.fc1(h))
        h2 = torch.tanh(self.fc2(h1)) + h1        # skip connection
        h3 = torch.tanh(self.fc3(h2)) + h2         # skip connection
        raw = self.fc_out(h3)                       # (B, N, 10)

        # ── State outputs with adaptive scaling [7] ───────────────────
        states = torch.cat([
            raw[..., 0:1] * self.xy_scale,           # x
            raw[..., 1:2] * self.xy_scale,           # y
            raw[..., 2:3] * self.z_scale,            # z
            raw[..., 3:4] * self.v_scale,            # vx
            raw[..., 4:5] * self.v_scale,            # vy
            raw[..., 5:6] * self.v_scale,            # vz
        ], dim=-1)

        # ── Control outputs with physical constraints [5] ─────────────
        ctrls = torch.cat([
            self.theta_max * torch.tanh(raw[..., 6:7]),     # θ ∈ [-θ_max, θ_max]
            self.phi_max * torch.tanh(raw[..., 7:8]),        # φ ∈ [-φ_max, φ_max]
            self.F_min + (self.F_max - self.F_min) * torch.sigmoid(raw[..., 8:9]),  # F ∈ [F_min, F_max]
        ], dim=-1)

        # ── Flight time ───────────────────────────────────────────────
        T = self.params.T_min + (self.params.T_max - self.params.T_min) * torch.sigmoid(raw[..., 9:10])

        return states, ctrls, T


# ═══════════════════════════════════════════════════════════════════════════
#  4. PINN LOSS (exact autograd)
# ═══════════════════════════════════════════════════════════════════════════

def compute_pinn_loss(model: PINNNetwork, task_batch: torch.Tensor,
                      cfg: TrainConfig, params: UAVParams,
                      device: torch.device) -> Dict[str, torch.Tensor]:
    """
    Full PINN loss with exact autograd for PDE residual.

    [13] PDE: dS/dτ = T · f(S, u)  computed via torch.autograd.grad
    [3]  Horizontal velocity penalty at landing
    [4]  Ground collision penalty
    [9]  Explicit IC at τ=0
    [15] Asymmetric drag
    """
    B, N = task_batch.shape[0], cfg.n_colloc

    # ══════════════════════════════════════════════════════════════════
    # A) PDE COLLOCATION LOSS — exact autograd [13]
    # ══════════════════════════════════════════════════════════════════
    tau = torch.rand(B, N, 1, device=device, requires_grad=True)
    task_exp = task_batch.unsqueeze(1).expand(-1, N, -1)  # (B, N, 4)

    states, ctrls, T_pred = model(tau, task_exp)

    # Exact per-channel gradient dS_i/dτ via autograd.
    # We compute each channel separately because states has 6 outputs
    # but tau has 1 input — a single grad call would sum over channels.
    d_states_list = []
    for i in range(6):
        g = torch.autograd.grad(
            outputs=states[..., i:i+1],
            inputs=tau,
            grad_outputs=torch.ones_like(tau),
            create_graph=True,
            retain_graph=True
        )[0]
        d_states_list.append(g)
    d_states = torch.cat(d_states_list, dim=-1)  # (B, N, 6)

    # Extract state and control components
    vx, vy, vz = states[..., 3:4], states[..., 4:5], states[..., 5:6]
    z = states[..., 2:3]
    theta, phi, F = ctrls[..., 0:1], ctrls[..., 1:2], ctrls[..., 2:3]
    vwx, vwy = task_exp[..., 2:3], task_exp[..., 3:4]

    # Effective thrust [6]
    if params.enable_ground_effect:
        z_safe = torch.clamp(z, min=0.1)
        ge = 1.0 + (params.R_rotor / (4.0 * z_safe)) ** 2
        F_eff = F * ge
    else:
        F_eff = F

    # UAV dynamics — PDF Eq.(1) with asymmetric drag [15], yaw=0 [2]
    ax = (F_eff * torch.sin(theta) * torch.cos(phi)
          - params.k_dx * (vx - vwx)) / params.m
    ay = (-F_eff * torch.sin(phi)
          - params.k_dy * (vy - vwy)) / params.m
    az = (F_eff * torch.cos(theta) * torch.cos(phi)
          - params.m * params.g
          - params.k_dz * vz) / params.m

    f_physics = torch.cat([vx, vy, vz, ax, ay, az], dim=-1)  # (B, N, 6)

    # PDE residual: dS/dτ should equal T * f(S, u)
    pde_residual = d_states - T_pred * f_physics
    loss_pde = torch.mean(pde_residual ** 2)

    # [4] Ground collision penalty
    z_colloc = states[..., 2:3]
    loss_ground = torch.mean(torch.relu(-z_colloc) ** 2)

    # ══════════════════════════════════════════════════════════════════
    # B) INITIAL CONDITION LOSS [9]
    # ══════════════════════════════════════════════════════════════════
    tau_0 = torch.zeros(B, 1, 1, device=device)
    task_0 = task_batch.unsqueeze(1)  # (B, 1, 4)
    s0, _, _ = model(tau_0, task_0)

    ic_target = torch.tensor([0., 0., params.z0, 0., 0., 0.], device=device)
    loss_ic = torch.mean((s0 - ic_target) ** 2)

    # ══════════════════════════════════════════════════════════════════
    # C) BOUNDARY CONDITION LOSS (landing)
    # ══════════════════════════════════════════════════════════════════
    tau_1 = torch.ones(B, 1, 1, device=device)
    s1, _, _ = model(tau_1, task_0)

    # Position: reach (xf, yf, 0)
    loss_bc_pos = torch.mean((s1[..., 0:2] - task_batch[:, 0:2].unsqueeze(1)) ** 2)
    loss_bc_alt = torch.mean(s1[..., 2:3] ** 2) * 3.0       # z → 0

    # [3] All velocities → 0 at landing
    loss_bc_vz = torch.mean(s1[..., 5:6] ** 2) * 2.0        # vertical velocity
    loss_bc_vxy = torch.mean(s1[..., 3:5] ** 2)              # horizontal velocity

    loss_bc = loss_bc_pos + loss_bc_alt + loss_bc_vz + cfg.w_vel_hz * loss_bc_vxy

    # ══════════════════════════════════════════════════════════════════
    # TOTAL LOSS
    # ══════════════════════════════════════════════════════════════════
    total = (cfg.w_pde * loss_pde
             + cfg.w_ic * loss_ic
             + cfg.w_bc * loss_bc
             + cfg.w_ground * loss_ground)

    return {
        'total': total,
        'pde': loss_pde,
        'ic': loss_ic,
        'bc': loss_bc,
        'ground': loss_ground,
    }


# ═══════════════════════════════════════════════════════════════════════════
#  5. TRAINING LOOP (Adam + L-BFGS)
# ═══════════════════════════════════════════════════════════════════════════

def train_model(model: PINNNetwork, params: UAVParams, cfg: TrainConfig,
                device: torch.device) -> Tuple[PINNNetwork, dict, dict]:
    """Two-phase training with per-epoch profiling for benchmarking."""

    history = {k: [] for k in ['pde', 'ic', 'bc', 'ground', 'total', 'epoch']}

    # ── Per-epoch profiling storage ───────────────────────────────────
    profiling = {
        'epoch': [], 'phase': [],
        'wall_time_s': [], 'epoch_time_ms': [],
        'loss_total': [], 'loss_pde': [], 'loss_ic': [],
        'loss_bc': [], 'loss_ground': [], 'learning_rate': [],
        'cpu_percent': [], 'ram_used_mb': [], 'ram_percent': [],
        'gpu_mem_allocated_mb': [], 'gpu_mem_reserved_mb': [],
        'gpu_utilization_pct': [],
    }

    def get_gpu_utilization():
        if device.type != 'cuda':
            return 0.0
        try:
            import subprocess
            r = subprocess.run(
                ['nvidia-smi', '--query-gpu=utilization.gpu',
                 '--format=csv,noheader,nounits', '-i', '0'],
                capture_output=True, text=True, timeout=2)
            return float(r.stdout.strip())
        except Exception:
            return 0.0

    last_gpu_util = [0.0]

    def record_epoch(epoch, phase, epoch_start, losses, lr, sample_gpu=False):
        now = time.time()
        profiling['epoch'].append(epoch)
        profiling['phase'].append(phase)
        profiling['wall_time_s'].append(round(now - t_start, 3))
        profiling['epoch_time_ms'].append(round((now - epoch_start) * 1000, 2))
        profiling['loss_total'].append(losses['total'].item())
        profiling['loss_pde'].append(losses['pde'].item())
        profiling['loss_ic'].append(losses['ic'].item())
        profiling['loss_bc'].append(losses['bc'].item())
        profiling['loss_ground'].append(losses['ground'].item())
        profiling['learning_rate'].append(lr)
        if HAS_PSUTIL:
            proc = psutil.Process(os.getpid())
            profiling['cpu_percent'].append(psutil.cpu_percent(interval=None))
            profiling['ram_used_mb'].append(round(proc.memory_info().rss / 1e6, 1))
            profiling['ram_percent'].append(round(psutil.virtual_memory().percent, 1))
        else:
            profiling['cpu_percent'].append(0.0)
            profiling['ram_used_mb'].append(0.0)
            profiling['ram_percent'].append(0.0)
        if device.type == 'cuda':
            profiling['gpu_mem_allocated_mb'].append(round(torch.cuda.memory_allocated() / 1e6, 1))
            profiling['gpu_mem_reserved_mb'].append(round(torch.cuda.memory_reserved() / 1e6, 1))
        else:
            profiling['gpu_mem_allocated_mb'].append(0.0)
            profiling['gpu_mem_reserved_mb'].append(0.0)
        if sample_gpu:
            last_gpu_util[0] = get_gpu_utilization()
        profiling['gpu_utilization_pct'].append(last_gpu_util[0])

    model.to(device)
    model.train()

    def sample_task(batch_size):
        task = torch.empty(batch_size, 4, device=device)
        task[:, 0:2].uniform_(-cfg.target_xy_range, cfg.target_xy_range)
        task[:, 2:4].uniform_(-cfg.wind_range, cfg.wind_range)
        return task

    def log_losses(losses, epoch):
        for k in ['pde', 'ic', 'bc', 'ground', 'total']:
            history[k].append(losses[k].item())
        history['epoch'].append(epoch)

    if HAS_PSUTIL:
        psutil.cpu_percent(interval=None)  # prime the measurement

    # ── Phase 1: Adam with cosine annealing [8] ──────────────────────
    optimizer = torch.optim.Adam(model.parameters(), lr=cfg.lr)
    scheduler = torch.optim.lr_scheduler.CosineAnnealingLR(
        optimizer, T_max=cfg.adam_epochs, eta_min=1e-5
    )

    print("=" * 70)
    print("  Phase 1: Adam Global Search")
    print("=" * 70)

    t_start = time.time()

    for epoch in range(1, cfg.adam_epochs + 1):
        epoch_t0 = time.time()
        task = sample_task(cfg.batch_size)

        optimizer.zero_grad()
        losses = compute_pinn_loss(model, task, cfg, params, device)
        losses['total'].backward()
        nn.utils.clip_grad_norm_(model.parameters(), max_norm=5.0)
        optimizer.step()
        scheduler.step()

        lr_now = optimizer.param_groups[0]['lr']
        sample_gpu = (epoch % 100 == 0 or epoch <= 2)
        record_epoch(epoch, 'adam', epoch_t0, losses, lr_now, sample_gpu)

        if epoch % cfg.print_every == 0 or epoch == 1:
            elapsed = time.time() - t_start
            print(f"  Epoch {epoch:5d} | Total: {losses['total'].item():9.4f} | "
                  f"PDE: {losses['pde'].item():.4f} | "
                  f"IC: {losses['ic'].item():.4f} | "
                  f"BC: {losses['bc'].item():.4f} | "
                  f"Gnd: {losses['ground'].item():.6f} | "
                  f"LR: {lr_now:.6f} | {elapsed:.1f}s")
            log_losses(losses, epoch)

    adam_time = time.time() - t_start
    print(f"\n  Adam complete in {adam_time:.1f}s\n")

    # ── Phase 2: L-BFGS precision tuning ─────────────────────────────
    print("=" * 70)
    print("  Phase 2: L-BFGS Precision Tuning")
    print("=" * 70)

    lbfgs_optimizer = torch.optim.LBFGS(
        model.parameters(), max_iter=20, history_size=50,
        line_search_fn='strong_wolfe')

    t_lbfgs = time.time()
    lbfgs_macro_steps = cfg.lbfgs_epochs // 20

    for step in range(1, lbfgs_macro_steps + 1):
        epoch_t0 = time.time()
        if step % 3 == 1:
            task_l = sample_task(cfg.batch_size)

        def closure():
            lbfgs_optimizer.zero_grad()
            l = compute_pinn_loss(model, task_l, cfg, params, device)
            l['total'].backward()
            return l['total']

        lbfgs_optimizer.step(closure)
        current_epoch = cfg.adam_epochs + step * 20

        eval_task = sample_task(cfg.batch_size)
        with torch.enable_grad():
            losses = compute_pinn_loss(model, eval_task, cfg, params, device)
        record_epoch(current_epoch, 'lbfgs', epoch_t0, losses, 0.0, step % 5 == 0)

        if step % 5 == 0 or step == 1:
            elapsed = time.time() - t_lbfgs
            print(f"  L-BFGS step {step:4d} (epoch ~{current_epoch}) | "
                  f"Total: {losses['total'].item():9.4f} | "
                  f"PDE: {losses['pde'].item():.4f} | "
                  f"IC: {losses['ic'].item():.4f} | "
                  f"BC: {losses['bc'].item():.4f} | {elapsed:.1f}s")
            log_losses(losses, current_epoch)

    lbfgs_time = time.time() - t_lbfgs
    total_time = time.time() - t_start
    print(f"\n  L-BFGS complete in {lbfgs_time:.1f}s")
    print(f"  Total training time: {total_time:.1f}s\n")

    return model, history, profiling


# ═══════════════════════════════════════════════════════════════════════════
#  6. RK4 NUMERICAL VERIFICATION [14]
# ═══════════════════════════════════════════════════════════════════════════

@torch.no_grad()
def rk4_verify(model: PINNNetwork, params: UAVParams,
               target: List[float], wind: List[float],
               device: torch.device, n_steps: int = 500) -> dict:
    """Independent RK4 forward integration using PINN-predicted controls."""
    model.eval()

    task_np = np.array([target[0], target[1], wind[0], wind[1]])
    task_t = torch.tensor(task_np, dtype=torch.float32, device=device).unsqueeze(0).unsqueeze(0)

    # Get predicted flight time
    tau_mid = torch.tensor([[[0.5]]], device=device)
    _, _, T_pred = model(tau_mid, task_t)
    T_val = T_pred[0, 0, 0].item()

    dt = T_val / n_steps
    state = np.array([0., 0., params.z0, 0., 0., 0.])
    rk4_traj = [state.copy()]

    def dynamics(s, tau_v):
        tau_t = torch.tensor([[[tau_v]]], device=device, dtype=torch.float32)
        _, ctrls, _ = model(tau_t, task_t)
        c = ctrls[0, 0].cpu().numpy()
        vx, vy, vz = s[3], s[4], s[5]
        theta, phi, F = c[0], c[1], c[2]
        ax_val = (F * math.sin(theta) * math.cos(phi) - params.k_dx * (vx - wind[0])) / params.m
        ay_val = (-F * math.sin(phi) - params.k_dy * (vy - wind[1])) / params.m
        az_val = (F * math.cos(theta) * math.cos(phi) - params.m * params.g - params.k_dz * vz) / params.m
        return np.array([vx, vy, vz, ax_val, ay_val, az_val])

    for i in range(n_steps):
        tau_v = i / n_steps
        dtau = 1.0 / n_steps
        k1 = dynamics(state, tau_v)
        k2 = dynamics(state + 0.5 * dt * k1, tau_v + 0.5 * dtau)
        k3 = dynamics(state + 0.5 * dt * k2, tau_v + 0.5 * dtau)
        k4 = dynamics(state + dt * k3, tau_v + dtau)
        state = state + (dt / 6.0) * (k1 + 2*k2 + 2*k3 + k4)
        rk4_traj.append(state.copy())

    rk4_traj = np.array(rk4_traj)

    # PINN trajectory for comparison
    n_pts = n_steps + 1
    tau_lin = torch.linspace(0, 1, n_pts, device=device).view(1, n_pts, 1)
    task_exp = task_t.expand(-1, n_pts, -1)
    pinn_states, pinn_ctrls, _ = model(tau_lin, task_exp)
    pinn_traj = pinn_states.squeeze(0).cpu().numpy()
    pinn_ctrl_arr = pinn_ctrls.squeeze(0).cpu().numpy()

    pos_err = np.sqrt(np.sum((rk4_traj[:, :3] - pinn_traj[:, :3]) ** 2, axis=1))

    return {
        'rk4_traj': rk4_traj,
        'pinn_traj': pinn_traj,
        'pinn_ctrls': pinn_ctrl_arr,
        'T': T_val,
        'pos_error': pos_err,
        'max_error': pos_err.max(),
        'mean_error': pos_err.mean(),
        'final_rk4': rk4_traj[-1],
        'final_pinn': pinn_traj[-1],
    }


# ═══════════════════════════════════════════════════════════════════════════
#  7. VISUALIZATION [12] — identical 6-panel dashboard
# ═══════════════════════════════════════════════════════════════════════════

C_BG = '#0A0A0F'; C_PANEL = '#12131A'; C_TEXT = '#D8D8DC'; C_GRID = '#1E2030'
# Metallic red accent palette
C_RED1  = '#FF2D2D'   # primary metallic red
C_RED2  = '#CC1F1F'   # darker red
C_RED3  = '#FF6B6B'   # lighter red / salmon
C_SILVER = '#A8B0BC'  # metallic silver
C_STEEL  = '#6B7A8D'  # steel gray
C_GOLD   = '#C9A84C'  # muted gold accent
C_EMBER  = '#FF4D00'  # ember orange-red


def _style_ax(ax, title="", xlabel="", ylabel=""):
    ax.set_facecolor(C_PANEL)
    ax.set_title(title, color=C_TEXT, fontsize=11, fontweight='bold', pad=12)
    ax.set_xlabel(xlabel, color=C_SILVER, fontsize=9, labelpad=6)
    ax.set_ylabel(ylabel, color=C_SILVER, fontsize=9, labelpad=6)
    ax.tick_params(colors=C_STEEL, labelsize=8)
    ax.grid(True, color=C_GRID, alpha=0.5, lw=0.4)
    for sp in ax.spines.values():
        sp.set_color(C_GRID)


def evaluate_and_visualize(model: PINNNetwork, params: UAVParams,
                           cfg: TrainConfig, history: dict,
                           device: torch.device):
    """6-panel dashboard + multi-scenario 3D plot. Identical layout to NumPy v2."""
    model.eval()

    scenarios = [
        {'target': [10.0, -5.0], 'wind': [2.0, -1.0], 'label': 'Crosswind'},
        {'target': [-6.0,  4.0], 'wind': [-1.5, 2.0], 'label': 'Reverse Wind'},
        {'target': [ 3.0,  3.0], 'wind': [0.0,  0.0], 'label': 'No Wind'},
        {'target': [ 8.0,  8.0], 'wind': [3.0,  3.0], 'label': 'Strong Diag.'},
    ]

    primary = scenarios[0]
    result = rk4_verify(model, params, primary['target'], primary['wind'], device)

    T_val = result['T']
    n_pts = result['pinn_traj'].shape[0]
    time_arr = np.linspace(0, T_val, n_pts)
    pinn_st = result['pinn_traj']
    pinn_ct = result['pinn_ctrls']
    rk4_st = result['rk4_traj']

    # ── Terminal report ───────────────────────────────────────────────
    print("=" * 70)
    print("  EVALUATION RESULTS")
    print("=" * 70)
    for i, sc in enumerate(scenarios):
        r = rk4_verify(model, params, sc['target'], sc['wind'], device)
        f = r['final_pinn']; t = sc['target']
        pe = math.sqrt((f[0]-t[0])**2 + (f[1]-t[1])**2 + f[2]**2)
        ve = math.sqrt(f[3]**2 + f[4]**2 + f[5]**2)
        print(f"\n  Scenario {i+1}: {sc['label']}")
        print(f"    Target: [{t[0]:.1f}, {t[1]:.1f}]  "
              f"Wind: [{sc['wind'][0]:.1f}, {sc['wind'][1]:.1f}] m/s")
        print(f"    Final pos: [{f[0]:.3f}, {f[1]:.3f}, {f[2]:.3f}]")
        print(f"    Position err: {pe:.4f} m | Velocity err: {ve:.4f} m/s | T: {r['T']:.2f}s")
        print(f"    RK4 max deviation: {r['max_error']:.4f} m | "
              f"mean: {r['mean_error']:.4f} m")
    print("=" * 70)

    # ══════════════════════════════════════════════════════════════════
    #  6-PANEL DASHBOARD — Dark + Metallic Red
    # ══════════════════════════════════════════════════════════════════
    fig = plt.figure(figsize=(22, 14), facecolor=C_BG)
    fig.suptitle('UAV PINN Precision Landing — PyTorch + CUDA',
                 color=C_RED1, fontsize=18, fontweight='bold', y=0.98,
                 fontfamily='monospace')
    gs = gridspec.GridSpec(2, 3, figure=fig, hspace=0.42, wspace=0.35,
                           left=0.06, right=0.96, top=0.91, bottom=0.07)

    # ── Panel 1: 3D Trajectory + RK4 ─────────────────────────────────
    ax1 = fig.add_subplot(gs[0, 0], projection='3d')
    ax1.set_facecolor(C_PANEL)
    for p in [ax1.xaxis.pane, ax1.yaxis.pane, ax1.zaxis.pane]:
        p.fill = False; p.set_edgecolor(C_GRID)
    ax1.plot(pinn_st[:, 0], pinn_st[:, 1], np.maximum(pinn_st[:, 2], 0),
             color=C_RED1, lw=2.5, label='PINN')
    ax1.plot(rk4_st[:, 0], rk4_st[:, 1], np.maximum(rk4_st[:, 2], 0),
             color=C_SILVER, lw=1.5, ls='--', label='RK4', alpha=0.7)
    ax1.scatter([0], [0], [params.z0], c=C_GOLD, s=120, marker='o',
                edgecolors='white', lw=0.8, label='Start', zorder=10)
    ax1.scatter([primary['target'][0]], [primary['target'][1]], [0],
                c=C_RED3, s=150, marker='X', edgecolors='white', lw=0.8,
                label='Target', zorder=10)
    wx, wy = primary['wind']
    ax1.quiver(0, 0, 0, wx*2, wy*2, 0, color=C_STEEL,
               arrow_length_ratio=0.2, lw=2, label=f'Wind [{wx},{wy}]')
    ax1.set_title('3D Trajectory + RK4', color=C_TEXT, fontsize=11,
                  fontweight='bold', pad=14)
    ax1.set_xlabel('X (m)', color=C_STEEL, fontsize=8, labelpad=4)
    ax1.set_ylabel('Y (m)', color=C_STEEL, fontsize=8, labelpad=4)
    ax1.set_zlabel('Z (m)', color=C_STEEL, fontsize=8, labelpad=4)
    ax1.tick_params(colors=C_STEEL, labelsize=7)
    ax1.legend(fontsize=7, facecolor=C_PANEL, edgecolor=C_GRID,
               labelcolor=C_TEXT, loc='upper left')

    # ── Panel 2: Velocity ─────────────────────────────────────────────
    ax2 = fig.add_subplot(gs[0, 1])
    _style_ax(ax2, 'Velocity Profile', 'Time (s)', 'Velocity (m/s)')
    ax2.plot(time_arr, pinn_st[:, 3], color=C_RED1, lw=2, label='vₓ')
    ax2.plot(time_arr, pinn_st[:, 4], color=C_SILVER, lw=2, label='vᵧ')
    ax2.plot(time_arr, pinn_st[:, 5], color=C_GOLD, lw=2, label='v_z')
    ax2.axhline(0, color=C_STEEL, ls='--', alpha=0.3, lw=0.8)
    ax2.axvline(T_val, color=C_RED3, ls=':', alpha=0.5, lw=1)
    ax2.annotate('T', xy=(T_val*0.97, ax2.get_ylim()[1]*0.85),
                 color=C_RED3, fontsize=8, ha='right')
    ax2.legend(fontsize=9, facecolor=C_PANEL, edgecolor=C_GRID,
               labelcolor=C_TEXT, loc='upper right')

    # ── Panel 3: Controls (NO twinx overlap — separate thrust & angle)
    ax3 = fig.add_subplot(gs[0, 2])
    _style_ax(ax3, 'Control Commands', 'Time (s)', 'Thrust F (N)')
    ax3.plot(time_arr, pinn_ct[:, 2], color=C_RED1, lw=2.5, label='Thrust F')
    ax3.axhline(params.m * params.g, color=C_STEEL, ls=':', alpha=0.4, lw=0.8)
    ax3.text(time_arr[5], params.m*params.g + 0.15, 'mg',
             color=C_STEEL, fontsize=7, alpha=0.6)
    # Separate twin axis with clear spacing
    ax3r = ax3.twinx()
    ax3r.spines['right'].set_color(C_SILVER)
    ax3r.plot(time_arr, np.degrees(pinn_ct[:, 0]), color=C_GOLD,
              ls='-.', lw=1.8, label='Pitch θ')
    ax3r.plot(time_arr, np.degrees(pinn_ct[:, 1]), color=C_STEEL,
              ls=':', lw=1.8, label='Roll φ')
    ax3r.set_ylabel('Angle (°)', color=C_SILVER, fontsize=9, labelpad=8)
    ax3r.tick_params(colors=C_STEEL, labelsize=8)
    # Combined legend positioned to avoid overlap
    lines_3 = ax3.get_lines() + ax3r.get_lines()
    labels_3 = [l.get_label() for l in lines_3]
    ax3.legend(lines_3, labels_3, fontsize=7, facecolor=C_PANEL,
               edgecolor=C_GRID, labelcolor=C_TEXT, loc='upper left')

    # ── Panel 4: Altitude + ground safety ─────────────────────────────
    ax4 = fig.add_subplot(gs[1, 0])
    _style_ax(ax4, 'Altitude & Ground Safety', 'Time (s)', 'Z (m)')
    ax4.fill_between(time_arr, 0, np.maximum(pinn_st[:, 2], 0),
                     color=C_RED2, alpha=0.08)
    ax4.plot(time_arr, pinn_st[:, 2], color=C_RED1, lw=2.5, label='PINN z')
    ax4.plot(time_arr, rk4_st[:, 2], color=C_SILVER, lw=1.5, ls='--',
             label='RK4 z', alpha=0.7)
    ax4.axhline(0, color=C_RED3, ls='-', alpha=0.5, lw=1.5, label='Ground')
    ax4.legend(fontsize=8, facecolor=C_PANEL, edgecolor=C_GRID,
               labelcolor=C_TEXT, loc='upper right')
    mn = min(pinn_st[:, 2].min(), rk4_st[:, 2].min())
    if mn < -0.1:
        ax4.annotate('GROUND BREACH', xy=(time_arr[np.argmin(pinn_st[:, 2])], mn),
                     color=C_RED1, fontsize=9, fontweight='bold')

    # ── Panel 5: Training loss ────────────────────────────────────────
    ax5 = fig.add_subplot(gs[1, 1])
    _style_ax(ax5, 'Training Loss Curves', 'Epoch', 'Loss (log)')
    epochs = history['epoch']
    loss_styles = [
        ('pde', C_RED1,   'o', 'PDE'),
        ('ic',  C_GOLD,   's', 'IC'),
        ('bc',  C_SILVER, '^', 'BC'),
        ('ground', C_RED3, 'd', 'GROUND'),
    ]
    for key, c, mk, lbl in loss_styles:
        vals = np.maximum(np.array(history[key]), 1e-12)
        ax5.semilogy(epochs, vals, color=c, lw=2, label=lbl, marker=mk, ms=3,
                     markevery=max(1, len(epochs) // 10))
    if cfg.adam_epochs < max(epochs):
        ax5.axvline(cfg.adam_epochs, color=C_STEEL, ls='--', alpha=0.4, lw=1)
        ax5.annotate('Adam → L-BFGS', xy=(cfg.adam_epochs, max(history['pde'])*0.5),
                     color=C_STEEL, fontsize=7, ha='right')
    ax5.legend(fontsize=8, facecolor=C_PANEL, edgecolor=C_GRID,
               labelcolor=C_TEXT, loc='upper right')

    # ── Panel 6: Multi-scenario bar chart ─────────────────────────────
    ax6 = fig.add_subplot(gs[1, 2])
    _style_ax(ax6, 'Multi-Scenario Accuracy', '', 'Error')
    names, pos_errs, vel_errs = [], [], []
    for sc in scenarios:
        r = rk4_verify(model, params, sc['target'], sc['wind'], device)
        f = r['final_pinn']; t = sc['target']
        names.append(sc['label'])
        pos_errs.append(math.sqrt((f[0]-t[0])**2 + (f[1]-t[1])**2 + f[2]**2))
        vel_errs.append(math.sqrt(f[3]**2 + f[4]**2 + f[5]**2))

    x_b = np.arange(len(scenarios)); w = 0.30
    b1 = ax6.bar(x_b - w/2 - 0.02, pos_errs, w, color=C_RED1, alpha=0.85,
                 label='Pos (m)', edgecolor=C_RED2, lw=0.8)
    b2 = ax6.bar(x_b + w/2 + 0.02, vel_errs, w, color=C_SILVER, alpha=0.75,
                 label='Vel (m/s)', edgecolor=C_STEEL, lw=0.8)
    ax6.set_xticks(x_b)
    ax6.set_xticklabels(names, rotation=20, ha='right', fontsize=7, color=C_TEXT)
    ax6.legend(fontsize=8, facecolor=C_PANEL, edgecolor=C_GRID,
               labelcolor=C_TEXT, loc='upper left')
    # Value labels above bars
    for bar in list(b1) + list(b2):
        h = bar.get_height()
        ax6.text(bar.get_x() + bar.get_width()/2., h + max(pos_errs)*0.03,
                 f'{h:.2f}', ha='center', va='bottom', color=C_TEXT, fontsize=7)

    # ── Save dashboard ────────────────────────────────────────────────
    out_path = "pinn_uav_dashboard_torch.png"
    fig.savefig(out_path, dpi=200, facecolor=C_BG, bbox_inches='tight')
    plt.close(fig)
    print(f"\n  Dashboard saved -> {out_path}")

    # ── Multi-scenario 3D plot ────────────────────────────────────────
    fig2 = plt.figure(figsize=(10, 8), facecolor=C_BG)
    ax = fig2.add_subplot(111, projection='3d')
    ax.set_facecolor(C_PANEL)
    for p in [ax.xaxis.pane, ax.yaxis.pane, ax.zaxis.pane]:
        p.fill = False; p.set_edgecolor(C_GRID)

    clrs = [C_RED1, C_SILVER, C_GOLD, C_RED3]
    for sc, c in zip(scenarios, clrs):
        r = rk4_verify(model, params, sc['target'], sc['wind'], device)
        tr = r['pinn_traj']
        ax.plot(tr[:, 0], tr[:, 1], np.maximum(tr[:, 2], 0), lw=2.5,
                color=c, label=sc['label'])
        ax.scatter([sc['target'][0]], [sc['target'][1]], [0], marker='X', s=100,
                   c=c, edgecolors='white', lw=0.6, zorder=10)
    ax.scatter([0], [0], [params.z0], c=C_GOLD, s=120, marker='o',
               edgecolors='white', lw=0.8, label='Start')

    ax.set_title('Multi-Scenario Trajectories', color=C_TEXT, fontsize=14,
                 fontweight='bold', pad=18, fontfamily='monospace')
    ax.set_xlabel('X (m)', color=C_STEEL, fontsize=9)
    ax.set_ylabel('Y (m)', color=C_STEEL, fontsize=9)
    ax.set_zlabel('Z (m)', color=C_STEEL, fontsize=9)
    ax.tick_params(colors=C_STEEL, labelsize=8)
    ax.legend(fontsize=8, facecolor=C_PANEL, edgecolor=C_GRID, labelcolor=C_TEXT)
    fig2.savefig("pinn_multi_trajectory_torch.png", dpi=200, facecolor=C_BG,
                 bbox_inches='tight')
    plt.close(fig2)
    print("  Multi-scenario plot saved -> pinn_multi_trajectory_torch.png\n")

    return result


# ═══════════════════════════════════════════════════════════════════════════
#  8. XLSX EXPORT (for Python vs Julia benchmarking)
# ═══════════════════════════════════════════════════════════════════════════

def export_results_xlsx(profiling: dict, history: dict, eval_results: list,
                        params: UAVParams, cfg: TrainConfig,
                        device: torch.device, total_params: int,
                        filename: str = "pinn_benchmark_python.xlsx"):
    """Export all training + evaluation data to xlsx for cross-language benchmarking."""

    if not HAS_OPENPYXL:
        print("  [WARN] openpyxl not installed. Saving CSV fallback instead.")
        import csv
        csv_path = filename.replace('.xlsx', '_profiling.csv')
        with open(csv_path, 'w', newline='') as f:
            w = csv.writer(f)
            w.writerow(profiling.keys())
            w.writerows(zip(*profiling.values()))
        print(f"  CSV fallback saved -> {csv_path}")
        return

    wb = openpyxl.Workbook()

    # ── Styling ───────────────────────────────────────────────────────
    hdr_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    hdr_fill = PatternFill('solid', fgColor='2F5496')
    sub_fill = PatternFill('solid', fgColor='D6E4F0')
    sub_font = Font(name='Arial', bold=True, size=10)
    data_font = Font(name='Arial', size=10)
    num_fmt_2 = '0.00'
    num_fmt_4 = '0.0000'
    num_fmt_6 = '0.000000'
    num_fmt_ms = '0.00'
    thin_border = Border(
        bottom=Side(style='thin', color='D0D0D0'))

    def style_header(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 3, 22)

    # ══════════════ Sheet 1: Per-Epoch Profiling ══════════════════════
    ws1 = wb.active
    ws1.title = "Training Profiling"

    headers = [
        'Epoch', 'Phase', 'Wall Time (s)', 'Epoch Time (ms)',
        'Loss Total', 'Loss PDE', 'Loss IC', 'Loss BC', 'Loss Ground',
        'Learning Rate', 'CPU %', 'RAM Used (MB)', 'RAM %',
        'GPU Mem Alloc (MB)', 'GPU Mem Reserved (MB)', 'GPU Util %'
    ]
    for c, h in enumerate(headers, 1):
        ws1.cell(row=1, column=c, value=h)
    style_header(ws1, 1, len(headers))

    n_rows = len(profiling['epoch'])
    keys = ['epoch', 'phase', 'wall_time_s', 'epoch_time_ms',
            'loss_total', 'loss_pde', 'loss_ic', 'loss_bc', 'loss_ground',
            'learning_rate', 'cpu_percent', 'ram_used_mb', 'ram_percent',
            'gpu_mem_allocated_mb', 'gpu_mem_reserved_mb', 'gpu_utilization_pct']

    for r in range(n_rows):
        for c, k in enumerate(keys):
            val = profiling[k][r]
            cell = ws1.cell(row=r + 2, column=c + 1, value=val)
            cell.font = data_font
            cell.border = thin_border
            if k in ('loss_total', 'loss_pde', 'loss_ic', 'loss_bc'):
                cell.number_format = num_fmt_6
            elif k == 'loss_ground':
                cell.number_format = '0.00E+00'
            elif k in ('wall_time_s', 'epoch_time_ms'):
                cell.number_format = num_fmt_2
            elif k == 'learning_rate':
                cell.number_format = '0.00E+00'

    ws1.freeze_panes = 'A2'
    auto_width(ws1)

    # ══════════════ Sheet 2: Summary Statistics ═══════════════════════
    ws2 = wb.create_sheet("Summary")

    summary_data = [
        ('System Info', '', ''),
        ('Language', 'Python', ''),
        ('Framework', f'PyTorch {torch.__version__}', ''),
        ('Device', str(device), ''),
        ('GPU Name', torch.cuda.get_device_name(0) if device.type == 'cuda' else 'N/A', ''),
        ('CUDA Version', str(torch.version.cuda) if device.type == 'cuda' else 'N/A', ''),
        ('OS', platform.platform(), ''),
        ('Python Version', platform.python_version(), ''),
        ('', '', ''),
        ('Model Config', '', ''),
        ('Network Parameters', total_params, ''),
        ('Hidden Dim', 256, ''),
        ('Hidden Layers', 4, ''),
        ('Activation', 'Tanh', ''),
        ('Adam Epochs', cfg.adam_epochs, ''),
        ('L-BFGS Epochs', cfg.lbfgs_epochs, ''),
        ('Batch Size', cfg.batch_size, ''),
        ('Collocation Points', cfg.n_colloc, ''),
        ('Initial LR', cfg.lr, ''),
        ('', '', ''),
        ('Physics Params', '', ''),
        ('Mass (kg)', params.m, ''),
        ('Gravity (m/s²)', params.g, ''),
        ('Drag k_dx', params.k_dx, ''),
        ('Drag k_dy', params.k_dy, ''),
        ('Drag k_dz', params.k_dz, ''),
        ('Initial Altitude z₀ (m)', params.z0, ''),
        ('', '', ''),
        ('Training Performance', '', ''),
        ('Total Wall Time (s)', profiling['wall_time_s'][-1] if profiling['wall_time_s'] else 0, ''),
        ('Adam Phase Time (s)', '', ''),
        ('L-BFGS Phase Time (s)', '', ''),
        ('Avg Epoch Time (ms)', round(np.mean(profiling['epoch_time_ms']), 2) if profiling['epoch_time_ms'] else 0, ''),
        ('Median Epoch Time (ms)', round(np.median(profiling['epoch_time_ms']), 2) if profiling['epoch_time_ms'] else 0, ''),
        ('Min Epoch Time (ms)', round(np.min(profiling['epoch_time_ms']), 2) if profiling['epoch_time_ms'] else 0, ''),
        ('Max Epoch Time (ms)', round(np.max(profiling['epoch_time_ms']), 2) if profiling['epoch_time_ms'] else 0, ''),
        ('Final Loss Total', profiling['loss_total'][-1] if profiling['loss_total'] else 0, ''),
        ('Final Loss PDE', profiling['loss_pde'][-1] if profiling['loss_pde'] else 0, ''),
        ('Peak GPU Memory (MB)', round(torch.cuda.max_memory_allocated() / 1e6, 1) if device.type == 'cuda' else 0, ''),
        ('Peak RAM (MB)', max(profiling['ram_used_mb']) if profiling['ram_used_mb'] else 0, ''),
    ]

    # Compute adam/lbfgs split times
    adam_epochs_data = [i for i, p in enumerate(profiling['phase']) if p == 'adam']
    lbfgs_epochs_data = [i for i, p in enumerate(profiling['phase']) if p == 'lbfgs']
    if adam_epochs_data:
        adam_wall = profiling['wall_time_s'][adam_epochs_data[-1]]
        summary_data[30] = ('Adam Phase Time (s)', round(adam_wall, 2), '')
    if lbfgs_epochs_data:
        lbfgs_wall = profiling['wall_time_s'][lbfgs_epochs_data[-1]] - (adam_wall if adam_epochs_data else 0)
        summary_data[31] = ('L-BFGS Phase Time (s)', round(lbfgs_wall, 2), '')

    for r, (label, value, note) in enumerate(summary_data, 1):
        c1 = ws2.cell(row=r, column=1, value=label)
        c2 = ws2.cell(row=r, column=2, value=value)
        c3 = ws2.cell(row=r, column=3, value=note)
        if label in ('System Info', 'Model Config', 'Physics Params', 'Training Performance'):
            c1.font = sub_font
            c1.fill = sub_fill
            c2.fill = sub_fill
            c3.fill = sub_fill
        else:
            c1.font = data_font
            c2.font = data_font

    auto_width(ws2)

    # ══════════════ Sheet 3: Evaluation Results ═══════════════════════
    ws3 = wb.create_sheet("Evaluation Results")

    eval_headers = [
        'Scenario', 'Target X', 'Target Y', 'Wind Vx', 'Wind Vy',
        'Final X', 'Final Y', 'Final Z',
        'Final Vx', 'Final Vy', 'Final Vz',
        'Position Error (m)', 'Velocity Error (m/s)',
        'Flight Time T (s)',
        'RK4 Max Deviation (m)', 'RK4 Mean Deviation (m)'
    ]
    for c, h in enumerate(eval_headers, 1):
        ws3.cell(row=1, column=c, value=h)
    style_header(ws3, 1, len(eval_headers))

    for r, ev in enumerate(eval_results, 2):
        vals = [
            ev['label'], ev['target'][0], ev['target'][1],
            ev['wind'][0], ev['wind'][1],
            ev['final_pos'][0], ev['final_pos'][1], ev['final_pos'][2],
            ev['final_vel'][0], ev['final_vel'][1], ev['final_vel'][2],
            ev['pos_error'], ev['vel_error'], ev['T'],
            ev['rk4_max_dev'], ev['rk4_mean_dev']
        ]
        for c, v in enumerate(vals, 1):
            cell = ws3.cell(row=r, column=c, value=round(v, 6) if isinstance(v, float) else v)
            cell.font = data_font
            cell.border = thin_border
            if isinstance(v, float):
                cell.number_format = num_fmt_4

    ws3.freeze_panes = 'A2'
    auto_width(ws3)

    # ══════════════ Sheet 4: Loss History (sampled) ═══════════════════
    ws4 = wb.create_sheet("Loss History")
    loss_headers = ['Epoch', 'Total', 'PDE', 'IC', 'BC', 'Ground']
    for c, h in enumerate(loss_headers, 1):
        ws4.cell(row=1, column=c, value=h)
    style_header(ws4, 1, len(loss_headers))

    for r in range(len(history['epoch'])):
        row_data = [
            history['epoch'][r], history['total'][r], history['pde'][r],
            history['ic'][r], history['bc'][r], history['ground'][r]
        ]
        for c, v in enumerate(row_data, 1):
            cell = ws4.cell(row=r + 2, column=c, value=v)
            cell.font = data_font
            if isinstance(v, float):
                cell.number_format = num_fmt_6

    ws4.freeze_panes = 'A2'
    auto_width(ws4)

    wb.save(filename)
    print(f"  Benchmark xlsx saved -> {filename}")


# ═══════════════════════════════════════════════════════════════════════════
#  9. MAIN
# ═══════════════════════════════════════════════════════════════════════════

def set_seed(seed: int = 42):
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed_all(seed)
        torch.backends.cudnn.deterministic = True
        torch.backends.cudnn.benchmark = False


def main():
    parser = argparse.ArgumentParser(description="UAV PINN v2.0 — PyTorch + CUDA")
    parser.add_argument("--device", type=str, default="auto",
                        choices=["auto", "cpu", "cuda", "mps"])
    parser.add_argument("--epochs", type=int, default=None)
    parser.add_argument("--lbfgs-epochs", type=int, default=None)
    parser.add_argument("--batch-size", type=int, default=None)
    parser.add_argument("--seed", type=int, default=42)
    parser.add_argument("--xlsx", type=str, default="pinn_benchmark_python.xlsx",
                        help="Output xlsx filename")
    args, _ = parser.parse_known_args()

    set_seed(args.seed)

    params = UAVParams()
    cfg = TrainConfig()

    if args.epochs is not None:
        cfg.adam_epochs = args.epochs
    if args.lbfgs_epochs is not None:
        cfg.lbfgs_epochs = args.lbfgs_epochs
    if args.batch_size is not None:
        cfg.batch_size = args.batch_size

    print("\n" + "=" * 70)
    print("  UAV PINN Precision Landing Controller v2.0")
    print("  PyTorch + CUDA Implementation")
    print("=" * 70)

    device = select_device(args.device)

    print(f"  PyTorch version: {torch.__version__}")
    if device.type == 'cuda':
        print(f"  CUDA version: {torch.version.cuda}")
        print(f"  cuDNN version: {torch.backends.cudnn.version()}")

    if HAS_PSUTIL:
        print(f"  psutil: available (CPU/RAM profiling ON)")
    else:
        print(f"  psutil: NOT installed (pip install psutil for full profiling)")
    if HAS_OPENPYXL:
        print(f"  openpyxl: available (xlsx export ON)")
    else:
        print(f"  openpyxl: NOT installed (pip install openpyxl for xlsx export)")

    print(f"\n  Mass: {params.m} kg | g: {params.g} m/s^2")
    print(f"  Drag (x,y,z): ({params.k_dx}, {params.k_dy}, {params.k_dz})  [15]")
    print(f"  Start: (0, 0, {params.z0}) | Yaw: {params.yaw} deg (fixed)  [2]")
    print(f"  Angles: +/-{params.theta_max_deg} deg pitch, "
          f"+/-{params.phi_max_deg} deg roll  [5]")
    print(f"  Thrust: [{params.thrust_min_ratio * params.m * params.g:.1f}, "
          f"{params.thrust_max_ratio * params.m * params.g:.1f}] N  [5]")
    print(f"  Flight time: [{params.T_min}, {params.T_max}] s")
    print(f"  Ground effect: {'ON' if params.enable_ground_effect else 'OFF'}  [6]")
    print(f"  Training: Adam({cfg.adam_epochs}) + L-BFGS({cfg.lbfgs_epochs}) | "
          f"Batch: {cfg.batch_size} | Colloc: {cfg.n_colloc}")
    print("=" * 70)

    model = PINNNetwork(params, cfg)
    total_params = sum(p.numel() for p in model.parameters())
    print(f"\n  Network: 5 -> 256 -> 256 -> 256 -> 256 -> 10 (skip connections)")
    print(f"  Total parameters: {total_params:,}")
    print(f"  Float precision: float32\n")

    trained_model, history, profiling = train_model(model, params, cfg, device)

    # Save model checkpoint
    ckpt_path = "pinn_uav_v2_torch.pt"
    torch.save({
        'model_state_dict': trained_model.state_dict(),
        'params': params, 'config': cfg, 'history': history,
    }, ckpt_path)
    print(f"  Model checkpoint saved -> {ckpt_path}")

    evaluate_and_visualize(trained_model, params, cfg, history, device)

    # ── Collect evaluation results for xlsx ───────────────────────────
    scenarios = [
        {'target': [10.0, -5.0], 'wind': [2.0, -1.0], 'label': 'Crosswind'},
        {'target': [-6.0,  4.0], 'wind': [-1.5, 2.0], 'label': 'Reverse Wind'},
        {'target': [ 3.0,  3.0], 'wind': [0.0,  0.0], 'label': 'No Wind'},
        {'target': [ 8.0,  8.0], 'wind': [3.0,  3.0], 'label': 'Strong Diag.'},
    ]
    eval_results = []
    for sc in scenarios:
        r = rk4_verify(trained_model, params, sc['target'], sc['wind'], device)
        f = r['final_pinn']; t = sc['target']
        eval_results.append({
            'label': sc['label'],
            'target': sc['target'], 'wind': sc['wind'],
            'final_pos': [f[0], f[1], f[2]],
            'final_vel': [f[3], f[4], f[5]],
            'pos_error': math.sqrt((f[0]-t[0])**2 + (f[1]-t[1])**2 + f[2]**2),
            'vel_error': math.sqrt(f[3]**2 + f[4]**2 + f[5]**2),
            'T': r['T'],
            'rk4_max_dev': r['max_error'],
            'rk4_mean_dev': r['mean_error'],
        })

    # ── Export xlsx ───────────────────────────────────────────────────
    export_results_xlsx(profiling, history, eval_results, params, cfg,
                        device, total_params, args.xlsx)

    if device.type == 'cuda':
        print(f"  Peak GPU memory: {torch.cuda.max_memory_allocated() / 1e6:.1f} MB")

    print("  All done.\n")


if __name__ == "__main__":
    main()